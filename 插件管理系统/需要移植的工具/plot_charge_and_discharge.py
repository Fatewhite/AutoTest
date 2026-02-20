# -*- coding: utf-8 -*-
import xlsxwriter
import fnmatch
import time
import os
import re
from collections import defaultdict
import numpy as np
import matplotlib
import pandas as pd
matplotlib.use('TkAgg') # 或者 'Qt5Agg'
# import matplotlib.pyplot as plt
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl import load_workbook



print('*************************充放电逻辑选择********************')
print('*    1.充电                                               *')
print('*    2.放电                                               *')
print('***********************************************************')
mode_choice = int(input("\n请选择充电or放电逻辑： "))
front = int(input("\n请输入删除前无用行的个数： "))
back  = int(input("\n请输出删除后无用行的个数:  "))

if mode_choice ==1:
    file_name = '充电'
elif mode_choice == 2:
    file_name = '放电'
else:
    file_name = None

batvolt_list = []
chgvolt_list = []
battemp_list = []
socekf_list  = []
socekfsend_list = []
batcurrent_list = []
chargecurrent_list = []
time_list = []

for f_name in os.listdir('.'):
    # print(f_name)
    if fnmatch.fnmatch(f_name, '*.txt'):
        f= f_name
        # print(f)
with open(f,'r', encoding='utf-8') as file:
    data_exc = file.readlines()
    del data_exc[0:front]
    del data_exc[-back:-1]
    del data_exc[-1]
# print(len(data_exc))
print(f"剩余总行数为：{len(data_exc)}")
print("数据处理中，时间可能较长，请耐心等待...")


def parse_list_to_dict(data_list):
    """
    将包含时间戳和变量数据的列表转换为字典
    """
    data_dict = defaultdict(list)
    for line in data_list:
        # 匹配时间戳
        timestamp_match = re.match(r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}:\d{3})', line)
        # print(timestamp_match)
        if timestamp_match:
            timestamp = timestamp_match.group(1)
            # print(type(timestamp))
            data_dict['timestamp'].append(timestamp)
        # 匹配所有 **变量名 : 值 的模式
        pattern = r'\*\*(\w+(?:\([^)]+\))?)\s*:\s*([^\s*]+)'
        matches = re.findall(pattern, line)
        # print(matches)
        for variable_name, variable_value in matches:
            try:
                if '.' in variable_value:
                    variable_value = float(variable_value)
                else:
                    variable_value = int(variable_value)
            except ValueError:
                pass
            data_dict[variable_name].append(variable_value)
    return dict(data_dict)
temp_dict = parse_list_to_dict(data_exc)


batvolt_list = temp_dict['batVolt(mV)']
chgvolt_list =temp_dict['chgVolt(mV)']
battemp_list =temp_dict['batTemp']
socekf_list  =temp_dict['SOCEkf']
socekfsend_list =temp_dict['SOCEkfSend']
batcurrent_list =temp_dict['batCurrent']
chargecurrent_list =temp_dict['chargeCurrent']
time_list =temp_dict['timestamp']



# for i in range(0, len(data_exc)):
#     for j in range(0, len(data_exc[i])):
#         if data_exc[i][j:j+3] == ':**':
#             time_list.append(data_exc[i][0:j])
#         if data_exc[i][j:j + 11] == 'batVolt(mV)' and data_exc[i][j:j + 16].strip() == 'batVolt(mV)':
#             batvolt_list.append(int(data_exc[i][j + 17:j + 24].strip()))
#         if data_exc[i][j:j + 11] == 'chgVolt(mV)' and data_exc[i][j:j + 16].strip() == 'chgVolt(mV)':
#             chgvolt_list.append(int(data_exc[i][j + 17:j + 24].strip()))
#         if data_exc[i][j:j + 7] == 'batTemp' and data_exc[i][j:j + 16].strip() == 'batTemp':
#             battemp_list.append(int(data_exc[i][j + 17:j + 24].strip()))
#         if data_exc[i][j:j + 6] == 'SOCEkf' and data_exc[i][j:j + 16].strip() == 'SOCEkf':
#             socekf_list.append(int(data_exc[i][j + 17:j + 24].strip()))
#         if data_exc[i][j:j + 10] == 'SOCEkfSend' and data_exc[i][j:j + 16].strip() == 'SOCEkfSend':
#             socekfsend_list.append(int(data_exc[i][j + 17:j + 24].strip()))
#         if data_exc[i][j:j + 10] == 'batCurrent' and data_exc[i][j:j + 16].strip() == 'batCurrent':
#             batcurrent_list.append(int(data_exc[i][j + 17:j + 24].strip()))
#         if data_exc[i][j:j + 13] == 'chargeCurrent' and data_exc[i][j:j + 16].strip() == 'chargeCurrent':
#             chargecurrent_list.append(int(data_exc[i][j + 17:j + 24].strip()))

# print(batvolt_list)
# print(chargecurrent_lsit)
# print(battemp_list)
# for i in range(0,len(data_exc)):
    # print(data_exc[i],end='')

# def arg_extr(temp_list=[],var_name=''):
#     l = len(var_name)
#     result_list = []
#     for i in range(0,len(temp_list)):
#         for j in range(0,len(temp_list[i])):
#             if temp_list[i][j:j+l] == var_name and temp_list[i][j:j+16].strip()==var_name:
#                 result_list.append(int(temp_list[i][j+17:j+24].strip()))
#     return result_list
#
# batvolt_list = arg_extr(data_exc,'batVolt(mV)')
# chgvolt_list = arg_extr(data_exc,'chgVolt(mV)')
# battemp_list = arg_extr(data_exc,'batTemp')
# socekf_list = arg_extr(data_exc,'SOCEkf')
# socekfsend_list = arg_extr(data_exc,'SOCEkfSend')
# batcurrent_list = arg_extr(data_exc,'batCurrent')
# chargecurrent_lsit = arg_extr(data_exc,'chargeCurrent')

x_list = []
for i in range(0,len(chargecurrent_list)):
    x_list.append(i+1)

# print(len(chargecurrent_list))
print("数据处理完成，正在生成EXCEL文件")

# data = {'时间':time_list,'电池电压':batvolt_list,'充电电压':chgvolt_list,'电池温度':battemp_list,'SOCEkf':socekf_list,\
#         '电量百分比':socekfsend_list,'电池电流':batcurrent_list,'充电电流':chargecurrent_list}
# df = pd.DataFrame(data)
# #生成excel表格
# df.to_excel(f'{file_name}数据整理.xlsx',index=False)
# print("表格已生成,正在绘制折线图")

# x = np.linspace(0, 30, 100)
# y1 = np.sin(x)
# y2 = np.cos(x)
# plt.figure(figsize=(20, 6))
# plt.plot(x,y1)
# plt.plot(time_list, i_list,marker='o', label='电流', color='red', linewidth=2,ms='4')
# plt.plot(time_list, v_list, marker='o',label='电压', color='blue', linewidth=2,ms='4', linestyle='--')

#  创建Excel文件
workbook = xlsxwriter.Workbook(f'{file_name}数据处理结果.xlsx')
worksheet = workbook.add_worksheet('Data')

header_format = workbook.add_format({
    'bold': True,          # 加粗
    'align': 'center',     # 水平居中
    'valign': 'vcenter',   # 垂直居中（可选）
    'border': 1,          # 边框（1=细线）
})

#  写入列标题和数据
worksheet.write(0, 0, "行号",header_format)#.write(行号,列号,内容)
worksheet.write(0, 1, "时间",header_format)
worksheet.write(0, 2, "电池电压",header_format)
worksheet.write(0, 3, "充电电压",header_format)
worksheet.write(0, 4, "电池温度",header_format)
worksheet.write(0, 5, "SOCEkf",header_format)
worksheet.write(0, 6, "电量百分比",header_format)
worksheet.write(0, 7, "电池电流",header_format)
worksheet.write(0, 8, "充电电流",header_format)

# 写入数据（行号自动生成）
for row in range(len(x_list)):
    # worksheet.write(row+1, 0, row+1)  # 行号（从1开始）
    # worksheet.write(row+1, 1, list1[row])
    # worksheet.write(row+1, 2, list2[row])
    worksheet.write(row+1, 0, x_list[row])
    worksheet.write(row+1, 1, time_list[row])
    worksheet.write(row+1, 2, batvolt_list[row])
    worksheet.write(row+1, 3, chgvolt_list[row])
    worksheet.write(row+1, 4, battemp_list[row])
    worksheet.write(row+1, 5, socekf_list[row])
    worksheet.write(row+1, 6, socekfsend_list[row])
    worksheet.write(row+1, 7, batcurrent_list[row])
    worksheet.write(row+1, 8, chargecurrent_list[row])


#  创建双Y轴折线图
chart = workbook.add_chart({'type': 'line'})

# 添加第一个系列（左Y轴）
chart.add_series({
    'name':       '=Data!$C$1',  # 引用列标题
    'categories': '=Data!$A$2:$A${}'.format(len(batvolt_list)+1),  # X轴（行号）
    'values':     '=Data!$C$2:$C${}'.format(len(batvolt_list)+1),  # Y轴数据
    'y2_axis':    False,  # 使用左Y轴
    'line':       {'color': 'blue', 'width': 2},
})
# 添加第二个系列（右Y轴）
chart.add_series({
    'name':       '=Data!$E$1',
    'categories': '=Data!$A$2:$A${}'.format(len(battemp_list)+1),  # 共享X轴就可生成双y轴的图形
    'values':     '=Data!$E$2:$E${}'.format(len(battemp_list)+1),
    'y2_axis':    True,  # 使用右Y轴
    'line':       {'color': 'red', 'width': 2},
})
#  设置图表格式
chart.set_title({'name': '电池电压与电池温度',
                 'name_font': {
                     'size': 14,  # 字体大小（磅值）
                     'bold': True,  # 加粗
                     # 'color': 'red'  # 字体颜色（可选）
                 }
                 })
# chart.set_x_axis({'name': '时间'})
chart.set_y_axis({'name': '电池电压',
                  'min':11000,
                  })
chart.set_y2_axis({'name': '电池温度',
                   'min':100})
chart.set_size({'width': 800, 'height': 500})
chart.set_legend({
    'position': 'bottom',
    'x': 0.05,
    'y': 0.05,
    'font': {'size': 9},
    # 'layout': {'x': 0, 'y': 0,'width':2,'heigth':3}  # 可同时设置图例字体
})

#  插入图表到工作表
worksheet.insert_chart('K3', chart)

#第二个图
chart1 = workbook.add_chart({'type': 'line'})
chart1.add_series({
    'name':       '=Data!$C$1',  # 引用列标题
    'categories': '=Data!$A$2:$A${}'.format(len(batvolt_list)+1),  # X轴（行号）
    'values':     '=Data!$C$2:$C${}'.format(len(batvolt_list)+1),  # Y轴数据
    'y2_axis':    False,  # 使用左Y轴
    'line':       {'color': 'blue', 'width': 2},
})

# 添加第二个系列（右Y轴）
chart1.add_series({
    'name':       '=Data!$G$1',
    'categories': '=Data!$A$2:$A${}'.format(len(battemp_list)+1),  # 共享X轴
    'values':     '=Data!$G$2:$G${}'.format(len(battemp_list)+1),
    'y2_axis':    True,  # 使用右Y轴
    'line':       {'color': 'red', 'width': 2},
})
#  设置图表格式
chart1.set_title({'name': '电池电压与充电百分比',
                  'name_font': {
                      'size': 14,  # 字体大小（磅值）
                      'bold': True,  # 加粗
                      # 'color': 'red'  # 字体颜色（可选）
                  }
                  })
# chart1.set_x_axis({'name': '时间'})
chart1.set_y_axis({'name': '电池电压',
                  'min':11000,
                  })
chart1.set_y2_axis({'name': '充电百分比'})
#  插入图表到工作表
chart1.set_size({'width': 800, 'height': 500})
chart1.set_legend({
    'position': 'bottom',
    'x': 0.8,
    'y': 0.1,
    'font': {'size': 9}  # 可同时设置图例字体
})
worksheet.insert_chart('K29', chart1)

if mode_choice == 1:
    chart2 = workbook.add_chart({'type': 'line'})
    chart2.add_series({
        'name': '=Data!$C$1',  # 引用列标题
        'categories': '=Data!$A$2:$A${}'.format(len(batvolt_list) + 1),  # X轴（行号）
        'values': '=Data!$C$2:$C${}'.format(len(batvolt_list) + 1),  # Y轴数据
        'y2_axis': False,  # 使用左Y轴
        'line': {'color': 'blue', 'width': 2},
    })
    # 添加第二个系列（右Y轴）
    chart2.add_series({
        'name': '=Data!$I$1',
        'categories': '=Data!$A$2:$A${}'.format(len(battemp_list) + 1),  # 共享X轴
        'values': '=Data!$I$2:$I${}'.format(len(battemp_list) + 1),
        'y2_axis': True,  # 使用右Y轴
        'line': {'color': 'red', 'width': 2},
    })

    #  设置图表格式
    chart2.set_title({'name': '电池电压与充电电流',
                      'name_font': {
                          'size': 14,  # 字体大小（磅值）
                          'bold': True,  # 加粗
                          # 'color': 'red'  # 字体颜色（可选）
                      }
                      })
    # chart2.set_x_axis({'name': '时间'})
    chart2.set_y_axis({'name': '电池电压',
                       'min': 11000,
                       })
    chart2.set_y2_axis({'name': '充电电流'})
    #  插入图表到工作表
    chart2.set_size({'width': 800, 'height': 500})
    chart2.set_legend({
        'position': 'bottom',
        'x': 0.8,
        'y': 0.1,
        'font': {'size': 9}  # 可同时设置图例字体
    })
    worksheet.insert_chart('K55', chart2)
elif mode_choice == 2:
    chart3 = workbook.add_chart({'type': 'line'})    #绘制折线图
    chart3.add_series({
        'name': '=Data!$C$1',  # 引用列标题
        'categories': '=Data!$A$2:$A${}'.format(len(batvolt_list) + 1),  # X轴（行号）
        'values': '=Data!$C$2:$C${}'.format(len(batvolt_list) + 1),  # Y轴数据
        'y2_axis': False,  # 使用左Y轴
        'line': {'color': 'blue', 'width': 2},
    })

    # 添加第二个系列（右Y轴）
    chart3.add_series({
        'name': '=Data!$H$1',
        'categories': '=Data!$A$2:$A${}'.format(len(battemp_list) + 1),  # 共享X轴
        'values': '=Data!$H$2:$H${}'.format(len(battemp_list) + 1),
        'y2_axis': True,  # 使用右Y轴
        'line': {'color': 'red', 'width': 2},
    })

    #  设置图表格式
    chart3.set_title({'name': '电池电压与电池电流',
                      'name_font': {
                          'size': 14,  # 字体大小（磅值）
                          'bold': True,  # 加粗
                          # 'color': 'red'  # 字体颜色（可选）
                      }
                      })
    # chart3.set_x_axis({'name': '时间'})
    chart3.set_y_axis({'name': '电池电压',
                       'min': 11000,
                       })
    chart3.set_y2_axis({'name': '电池电流',
                        'min':600,})
    #  插入图表到工作表
    chart3.set_size({'width': 800, 'height': 500})
    chart3.set_legend({
        'position': 'bottom',
        'x': 0.8,
        'y': 0.1,
        'font': {'size': 9}  # 可同时设置图例字体
    })
    worksheet.insert_chart('K55', chart3)
#设置B列和G列的列宽
worksheet.set_column('B:B', 24)
worksheet.set_column('G:G', 11)

# header_format = workbook.add_format({
#     'bold': True,          # 加粗
#     'align': 'center',     # 水平居中
#     'valign': 'vcenter',   # 垂直居中（可选）
#     'border': 1,          # 边框（1=细线）
# })

# 应用格式到第一行（A1:Z1）
# worksheet.set_row(0, None, header_format)  

# for col in range(0, 10):  # 0=A, 1=B, ..., 9=J
#     worksheet.write(0, col, f"Header{col+1}", header_format) 
#冻结第一行
worksheet.freeze_panes(1, 0)
#  保存文件
workbook.close()
print(f"处理完成，文件已保存为: {file_name}数据处理结果.xlsx")

# plt.ylim(0, None)
# if mode_choice ==2:
#     fig, ax1 = plt.subplots(figsize=(8, 5))
#     ax1.plot(x_list, batvolt_list, 'b-', label='batvolt')
#     ax1.set_xlabel('time')
#     ax1.set_ylabel('batvolt', color='b')
#     ax1.tick_params(axis='y', labelcolor='b')
#     # 创建第二个Y轴（共享X轴）
#     plt.ylim(0, None)
#     ax2 = ax1.twinx()
#     ax2.plot(x_list, batcurrent_list, 'r-', label='batcurrent')
#     ax2.set_ylabel('batcurrent', color='r')
#     ax2.tick_params(axis='y', labelcolor='r')
#
#     # 合并图例
#     lines1, labels1 = ax1.get_legend_handles_labels()
#     lines2, labels2 = ax2.get_legend_handles_labels()
#     ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left')
#     plt.ylim(0, None)
#
#     plt.grid(True) # 显示网格
#     plt.title('batcurrent and batvolt')
#     # plt.show()
#     plt.savefig('batcurrent_and_batvolt.png', dpi=300, bbox_inches='tight')
#     plt.show()
#     plt.close()
# elif mode_choice == 1:
#     fig, ax1 = plt.subplots(figsize=(8, 5))
#     ax1.plot(x_list, batvolt_list, 'b-', label='batvolt')
#     ax1.set_xlabel('time')
#     ax1.set_ylabel('batvolt', color='b')
#     ax1.tick_params(axis='y', labelcolor='b')
#
#     # 创建第二个Y轴（共享X轴）
#     plt.ylim(0, None)
#     ax2 = ax1.twinx()
#     ax2.plot(x_list, chargecurrent_list, 'r-', label='chargecurrent')
#     ax2.set_ylabel('chargecurrent', color='r')
#     ax2.tick_params(axis='y', labelcolor='r')
#
#     # 合并图例
#     lines1, labels1 = ax1.get_legend_handles_labels()
#     lines2, labels2 = ax2.get_legend_handles_labels()
#     ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left')
#     plt.ylim(0, None)
#
#     plt.grid(True)  # 显示网格
#     plt.title('chargecurrent and batvolt')
#     # plt.show()
#     plt.savefig('chargecurrent_and_batvolt.png', dpi=300, bbox_inches='tight')
#     plt.show()
#     plt.close()
# else:
#     pass
#
#
# # plt.ylim(0, None)
# fig, ax1 = plt.subplots(figsize=(8, 5))
# ax1.plot(x_list, batvolt_list, 'b-', label='batvolt')
# ax1.set_xlabel('time')
# ax1.set_ylabel('batvolt', color='b')
# ax1.tick_params(axis='y', labelcolor='b')
#
# # 创建第二个Y轴（共享X轴）
# plt.ylim(0, None)
# ax2 = ax1.twinx()
# ax2.plot(x_list, battemp_list, 'r-', label='batTemp')
# ax2.set_ylabel('batTemp', color='r')
# ax2.tick_params(axis='y', labelcolor='r')
#
# # 合并图例
# lines1, labels1 = ax1.get_legend_handles_labels()
# lines2, labels2 = ax2.get_legend_handles_labels()
# ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left')
# plt.ylim(0, None)
# plt.grid(True) # 显示网格
# plt.title('batTemp and batvolt')
# # plt.show()
# plt.savefig('batTemp_and_batvolt.png', dpi=300, bbox_inches='tight')
# plt.show()
# plt.close()
#
#
# # plt.ylim(0, None)
# fig, ax1 = plt.subplots(figsize=(8, 5))
# ax1.plot(x_list, batvolt_list, 'b-', label='batvolt')
# ax1.set_xlabel('time')
# ax1.set_ylabel('batvolt', color='b')
# ax1.tick_params(axis='y', labelcolor='b')
#
# # 创建第二个Y轴（共享X轴）
# plt.ylim(0, None)
# ax2 = ax1.twinx()
# ax2.plot(x_list, socekfsend_list, 'r-', label='SOCEkfSend')
# ax2.set_ylabel('SOCEkfSend', color='r')
# ax2.tick_params(axis='y', labelcolor='r')
#
# # 合并图例
# lines1, labels1 = ax1.get_legend_handles_labels()
# lines2, labels2 = ax2.get_legend_handles_labels()
# ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left')
#
# plt.ylim(0, None)
# plt.grid(True) # 显示网格
# plt.title('SOCEkfSend and batvolt')
# # plt.show()
# plt.savefig('SOCEkfSend_and_batvolt.png', dpi=300, bbox_inches='tight')
# plt.show()
# plt.close()

# wb = load_workbook('电池数据整理.xlsx')
# ws = wb.active

# img1 = Image('chargeCurrent_and_batvolt.png')
# img2 = Image('batTemp_and_batvolt.png')
# img3 = Image('SOCEkfSend_and_batvolt.png')
# ws.add_image(img1, 'J2')  # 图片左上角定位到  单元格
# ws.add_image(img2, 'J40')  # 图片左上角定位到  单元格
# ws.add_image(img3, 'J100')  # 图片左上角定位到  单元格

# 4. 保存 Excel
# wb.save('电池数据整理.xlsx')


# plt.figure(figsize=(20, 6))
# plt.plot(x_list, chargecurrent_list,marker='o', label='电流', color='red', linewidth=2,ms='4')
# plt.plot(x_list, batvolt_list, marker='o',label='电压', color='blue', linewidth=2,ms='4', linestyle='--')
# plt.show()



